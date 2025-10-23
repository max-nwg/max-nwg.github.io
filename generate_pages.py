import re
import sys
from pathlib import Path
from typing import Dict, Set

try:
	from openpyxl import load_workbook
except ImportError as e:
	print("Missing dependency: openpyxl. Install with: uv add openpyxl", file=sys.stderr)
	raise

try:
	from tqdm import tqdm
except ImportError as e:
	print("Missing dependency: tqdm. Install with: uv add tqdm", file=sys.stderr)
	raise

WORKSPACE_ROOT = Path(__file__).parent
DATA_XLSX = WORKSPACE_ROOT / "data.xlsx"
OUTPUT_DIR = WORKSPACE_ROOT / "output"

TEMPLATE_MAP = {
	"a": WORKSPACE_ROOT / "template_a.html",
	"b": WORKSPACE_ROOT / "template_b.html",
	"c": WORKSPACE_ROOT / "template_c.html",
	"d": WORKSPACE_ROOT / "template_d.html",
	"e": WORKSPACE_ROOT / "template_e.html",
	"f": WORKSPACE_ROOT / "template_f.html",
}

PLACEHOLDER_PATTERN = re.compile(r"\{\{\s*([a-zA-Z0-9_]+)\s*\}\}")



def format_cell_to_html(cell) -> str:
	"""
	Convert Excel cell content with formatting to HTML markup.
	Handles line breaks and bold formatting with support for mixed formatting.
	"""
	if cell.value is None:
		return ""
	
	# Check for rich text using _value attribute
	cell_value = cell._value if hasattr(cell, '_value') else cell.value
	
	# Check if it's a CellRichText object
	from openpyxl.cell.rich_text import CellRichText, TextBlock
	
	if isinstance(cell_value, CellRichText):
		html_parts = []
		for text_run in cell_value:
			if isinstance(text_run, TextBlock):
				text = text_run.text or ""
			else:
				text = str(text_run)
			
			if not text:
				continue
			
			# Handle line breaks
			text = text.replace('\n', '<br>')
			
			# Check for bold formatting
			is_bold = False
			if isinstance(text_run, TextBlock) and text_run.font:
				is_bold = text_run.font.b or text_run.font.bold or False
			
			# Apply bold formatting
			if is_bold:
				text = f"<strong>{text}</strong>"
			
			html_parts.append(text)
		
		return "".join(html_parts)
	
	# Fallback for regular cells
	text = str(cell.value)
	text = text.replace('\n', '<br>')
	
	if cell.font and (cell.font.b or cell.font.bold):
		text = f"<strong>{text}</strong>"
	
	return text


def read_workbook(path: Path):
	if not path.exists():
		raise FileNotFoundError(f"Excel file not found: {path}")
	# Use rich_text=True to preserve rich text formatting from Excel
	return load_workbook(filename=str(path), rich_text=True)


def extract_placeholders(template_text: str) -> Set[str]:
	return set(PLACEHOLDER_PATTERN.findall(template_text))


def render_template(template_text: str, values: Dict[str, str]) -> str:
	def replace(match: re.Match) -> str:
		key = match.group(1).strip()
		return str(values.get(key, ""))
	return PLACEHOLDER_PATTERN.sub(replace, template_text)


def sanitize_rel_path(rel_path: str) -> Path:
	# Ensure it is a POSIX-like path coming from Excel like /01/5404045500009/10/C555161
	# Strip leading/trailing whitespace and leading slashes, keep internal separators
	clean = (rel_path or "").strip()
	if clean.startswith("/"):
		clean = clean[1:]
	# Prevent path traversal
	parts = [p for p in clean.split("/") if p not in ("", ".", "..")]
	return Path(*parts)


def main() -> int:
	wb = read_workbook(DATA_XLSX)
	ws = wb.active

	# Expect: Column A = type (a-f), Column B = path, other columns are key names matching placeholders
	# Build header map from first row
	headers = []
	for cell in ws[1]:
		headers.append(str(cell.value).strip() if cell.value is not None else "")
	if len(headers) < 2:
		raise ValueError("Expected at least two columns: type and path")

	# Indices
	try:
		type_idx = 0
		path_idx = 1
	except ValueError as e:
		raise ValueError("Could not find required columns 'type' and 'path' in first two columns") from e

	# Count total rows first for progress bar
	total_rows = sum(1 for _ in ws.iter_rows(min_row=2, values_only=False))
	
	rows_processed = 0
	skipped_rows = 0
	
	# Create progress bar
	with tqdm(total=total_rows, desc="Generating pages", unit="page") as pbar:
		for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
			if row is None:
				pbar.update(1)
				continue
			
			# Read type and path
			type_cell = row[type_idx]
			type_value = (type_cell.value or "").strip().lower() if isinstance(type_cell.value, str) else str(type_cell.value or "").strip().lower()
			if not type_value:
				pbar.update(1)
				skipped_rows += 1
				continue
			
			template_path = TEMPLATE_MAP.get(type_value)
			if not template_path or not template_path.exists():
				pbar.set_description(f"Skipping row: unknown template type '{type_value}'")
				pbar.update(1)
				skipped_rows += 1
				continue

			path_cell = row[path_idx]
			path_value = (path_cell.value or "") if not isinstance(path_cell.value, (int, float)) else str(path_cell.value)
			rel_path = sanitize_rel_path(path_value)
			if rel_path == Path():
				pbar.set_description("Skipping row: empty path")
				pbar.update(1)
				skipped_rows += 1
				continue

			# Update progress bar description
			pbar.set_description(f"Processing {rel_path}")

			template_text = template_path.read_text(encoding="utf-8")
			placeholders = extract_placeholders(template_text)

			# Build value map from headers -> row values; keys must exactly match placeholders
			values: Dict[str, str] = {}
			for col_idx, header in enumerate(headers):
				if not header:
					continue
				if header in placeholders:
					# Get the actual cell object to preserve formatting
					cell = row[col_idx]
					val_str = format_cell_to_html(cell)
					values[header] = val_str

			# Render and write file
			output_dir = OUTPUT_DIR / rel_path
			# Ensure final directory ends at the last segment (e.g., .../C555161)
			output_dir.mkdir(parents=True, exist_ok=True)
			output_file = output_dir / "index.html"

			result_html = render_template(template_text, values)
			output_file.write_text(result_html, encoding="utf-8")
			rows_processed += 1
			
			# Update progress bar
			pbar.update(1)

	print(f"\nGenerated {rows_processed} pages into {OUTPUT_DIR}")
	if skipped_rows > 0:
		print(f"Skipped {skipped_rows} rows")
	return 0


if __name__ == "__main__":
	sys.exit(main())
